import argparse
import copy
import os
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
DESIGN_DIR = ROOT / "site/downloads/research-design"
FIXED_TIMESTAMP = datetime(2026, 8, 12)
FONT_NAME = "Malgun Gothic"
HEADER_FILL = "FF17365D"
ALT_FILL = "FFEAF2F8"
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFFFF")
BODY_FONT = Font(name=FONT_NAME, size=10, color="FF1F2937")
LINK_FONT = Font(name=FONT_NAME, size=10, color="FF0563C1", underline="single")
HEADER_BORDER = Border(
    right=Side(style="thin", color="FFFFFFFF"),
    bottom=Side(style="medium", color="FF8EA9C1"),
)
BODY_BORDER = Border(bottom=Side(style="thin", color="FFD9E2F3"))
PUBLIC_SITE_BASE = "https://beaten-to-it.github.io/paper/"
DENSE_PRINT_SHEETS = {"구성개념", "명제추적", "논문별기여", "사건코딩", "관계망"}
PRINT_TITLE_COLUMNS = {
    "구성개념": "A:A",
    "명제추적": "A:A",
    "논문별기여": "A:A",
    "사건코딩": "A:A",
    "관계망": "A:D",
}

EVENT_COLUMNS = (
    "event_code",
    "actor_role",
    "trigger",
    "threat",
    "role_complementarity",
    "activated_connectivity",
    "voice_safety",
    "voice_efficacy",
    "response_codes",
    "dominant_path",
    "substantive_handling",
    "problem_resolution",
    "rival_explanation",
    "negative_case",
    "dv1_decision",
    "dv1_primary_source",
    "dv1_corroboration",
    "dv2_decision",
    "dv2_primary_source",
    "dv2_corroboration",
    "dv3_decision",
    "dv3_primary_source",
    "dv3_corroboration",
    "dv4_decision",
    "dv4_primary_source",
    "dv4_corroboration",
    "divergence_vector",
    "divergence_event_decision",
    "divergence_aggregation",
    "divergence_uncertainty",
    "p6_use_decision",
    "p7_use_decision",
)
RELATION_TYPES = ("advice", "trust", "approval_exception", "risk_escalation")
RELATION_EVIDENCE_SUFFIXES = (
    "sender_report_delivery",
    "receipt_confirmation_evidence",
    "recipient_stance",
)
RELATION_COLUMNS = (
    "event_code",
    "respondent_code",
    "alter_code",
    "alter_role_category",
    *(f"{relation}_relation" for relation in RELATION_TYPES),
    "contact_frequency",
    "perceived_trust",
    *(
        f"{relation}_{suffix}"
        for relation in RELATION_TYPES
        for suffix in RELATION_EVIDENCE_SUFFIXES
    ),
)
ETHICS_SAFEGUARD_CONTRACT = (
    ("ETH01", "사용 전 gate", ("인터뷰와 관계망 조사 전에", "IRB", "기업 보안·법무 승인")),
    ("ETH02", "사용 전 gate", ("승인 전에는 파일럿 대상자에게 접촉하지 않고", "비승인 도구 사용기록")),
    ("ETH03", "사용 전 gate", ("승인 후에도", "규정위반 적발", "인사평가", "성과 감시")),
    ("ETH04", "연결코드와 내부자 safeguards", ("비응답 지명자", "역할 범주", "사건에만 귀속")),
    ("ETH05", "연결코드와 내부자 safeguards", ("연결코드 명부", "분리 암호화", "분석 종료 시 폐기")),
    ("ETH06", "연결코드와 내부자 safeguards", ("재직·협력 관계", "위치성·이해상충")),
    ("ETH07", "연결코드와 내부자 safeguards", ("지휘·평가 관계가 없는", "독립 모집 담당자")),
    ("ETH08", "연결코드와 내부자 safeguards", ("고용주에게 원자료를 제공하지 않는다", "동의서에 사전 명시")),
    ("ETH09", "연결코드와 내부자 safeguards", ("작은 팀의 관계망", "최소 셀 기준", "원응답을 열람하지 못한다")),
    ("ETH10", "중단 조건", ("IRB 또는 기업 보안·법무 승인 범위", "자료 수집을 중단")),
    ("ETH11", "중단 조건", ("평가·보복·재식별 위험", "질문을 즉시 건너뛰고", "철회")),
    ("ETH12", "중단 조건", ("실명, 고객명, 소스코드", "기록을 중지", "비식별 절차")),
    ("ETH13", "중단 조건", ("독립 모집, 원자료 비공유, 분리 보관", "자료 수집을 중단")),
    ("ETH14", "중단 조건", ("60분 예산", "관계 쌍을 늘리지 않고", "파일럿 부정 사례")),
)


@dataclass(frozen=True)
class Link:
    label: str
    target: str


def _public_link(path: str) -> str:
    return f"{PUBLIC_SITE_BASE}{path}"


def _read(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def _plain(text: str) -> str:
    text = re.sub(r"\[([^]]+)]\([^)]+\)", r"\1", text)
    text = text.replace("**", "").replace("`", "")
    return re.sub(r"\s+", " ", text).strip()


def _table_rows(text: str) -> list[list[str]]:
    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|") or not line.endswith("|"):
            continue
        cells = [cell.strip() for cell in line[1:-1].split("|")]
        if all(re.fullmatch(r":?-{3,}:?", cell) for cell in cells):
            continue
        rows.append(cells)
    return rows


def _heading_sections(text: str, level: int) -> list[tuple[str, str]]:
    matches = list(re.finditer(rf"^#{{{level}}} (.+)$", text, re.MULTILINE))
    return [
        (
            match.group(1).strip(),
            text[match.end() : matches[index + 1].start() if index + 1 < len(matches) else len(text)],
        )
        for index, match in enumerate(matches)
    ]


def _section(text: str, heading: str, level: int = 2) -> str:
    for found_heading, body in _heading_sections(text, level):
        if found_heading == heading:
            return body
    raise ValueError(f"missing Markdown section: {heading}")


def _bullet_text(section: str) -> str:
    return "\n".join(_plain(line[2:]) for line in section.splitlines() if line.strip().startswith("- "))


def _construct_rows() -> list[list[object]]:
    path = DESIGN_DIR / "construct-dictionary-v0.3.md"
    text = _read(path)
    fields = (
        "분석 수준",
        "정의",
        "포함",
        "제외",
        "관찰 가능한 근거",
        "시간적 위치",
        "가장 가까운 경쟁 구성개념",
        "귀속 규칙",
    )
    rows = []
    for heading, body in _heading_sections(text, 3):
        values = {cells[0]: _plain(cells[1]) for cells in _table_rows(body) if len(cells) == 2}
        if all(field in values for field in fields):
            rows.append([heading, *(values[field] for field in fields), Link("구성개념 사전", _public_link("downloads/research-design/construct-dictionary-v0.3.md"))])
    if len(rows) != 12:
        raise ValueError(f"expected 12 construct rows, found {len(rows)}")
    return rows


def _proposition_rows() -> list[list[object]]:
    overview = _read(DESIGN_DIR / "research-model-v0.3.md")
    traceability = _read(DESIGN_DIR / "proposition-traceability-v0.3.md")
    status_rows = {
        cells[0]: (_plain(cells[1]), _plain(cells[2]))
        for cells in _table_rows(_section(overview, "P1-P7 상태"))
        if len(cells) == 3 and re.fullmatch(r"P[1-7]", cells[0])
    }
    fields = (
        "직접 근거 출처",
        "통합 해석",
        "분석 수준",
        "자료원",
        "지지 패턴",
        "반증 패턴",
        "경쟁 설명",
        "연구 2 처분",
    )
    traced = {}
    for heading, body in _heading_sections(traceability, 3):
        match = re.match(r"(P[1-7])\s+—", heading)
        if not match:
            continue
        values = {cells[0]: _plain(cells[1]) for cells in _table_rows(body) if len(cells) == 2}
        if all(field in values for field in fields):
            traced[match.group(1)] = values
    rows = []
    for number in range(1, 8):
        proposition = f"P{number}"
        statement, status = status_rows[proposition]
        values = traced[proposition]
        rows.append(
            [
                proposition,
                statement,
                status,
                *(values[field] for field in fields),
                "미검토",
                Link("명제 추적표", _public_link("downloads/research-design/proposition-traceability-v0.3.md")),
            ]
        )
    return rows


def _contribution_rows() -> list[list[object]]:
    slugs = (
        "kemell-2025",
        "neumann-2026",
        "golgeci-2025",
        "battilana-casciaro-2012",
        "battilana-casciaro-2013",
    )
    rows = []
    for slug in slugs:
        path = ROOT / "site/downloads" / slug / "model-v0.3-contribution.md"
        text = _read(path)
        title = re.search(r"^# v0\.3 기여 카드 — (.+)$", text, re.MULTILINE)
        if not title:
            raise ValueError(f"missing contribution-card title: {path}")
        citation = title.group(1).strip()
        rows.append(
            [
                citation,
                "논문 직접 근거",
                _bullet_text(_section(text, "직접 근거")),
                _bullet_text(_section(text, "v0.3에서의 역할")),
                _bullet_text(_section(text, "연결되는 명제")),
                _bullet_text(_section(text, "검증하지 않은 것")),
                _bullet_text(_section(text, "현장자료 요구")),
                Link(f"{citation} 기여 카드", _public_link(f"downloads/{slug}/model-v0.3-contribution.md")),
            ]
        )
    return rows


def _section_items(text: str, heading: str) -> list[str]:
    items = []
    paragraph = []
    in_fence = False

    def flush_paragraph() -> None:
        if paragraph:
            text = _plain(" ".join(paragraph))
            items.append(text)
            items.extend(part for part in re.split(r"(?<=다\.)\s+", text) if part != text)
            paragraph.clear()

    for raw_line in _section(text, heading).splitlines():
        line = raw_line.strip()
        if line.startswith("```"):
            flush_paragraph()
            in_fence = not in_fence
        elif in_fence or line.startswith(("#", "|")):
            flush_paragraph()
        elif not line:
            flush_paragraph()
        elif line.startswith(("- ", "* ")):
            flush_paragraph()
            items.append(_plain(line[2:]))
        else:
            paragraph.append(line)
    flush_paragraph()
    return items


def _ethics_rows(text: str | None = None) -> list[list[object]]:
    target = _public_link("downloads/research-design/pilot-protocol-and-codingbook-v0.3.md")
    if text is None:
        text = _read(DESIGN_DIR / "pilot-protocol-and-codingbook-v0.3.md")
    items_by_section = {
        heading: _section_items(text, heading)
        for heading in {contract[1] for contract in ETHICS_SAFEGUARD_CONTRACT}
    }
    rows = []
    for safeguard_id, heading, required_terms in ETHICS_SAFEGUARD_CONTRACT:
        matches = [
            item
            for item in items_by_section[heading]
            if all(term in item for term in required_terms)
        ]
        if not matches:
            raise ValueError(f"missing {safeguard_id} safeguard in {heading}")
        rows.append([safeguard_id, min(matches, key=len), None, None, None, Link("파일럿 프로토콜·코딩북", target)])
    return rows


def _display_width(value: object) -> int:
    text = str(value)
    return sum(2 if ord(character) > 127 else 1 for character in text)


def _add_sheet(workbook: Workbook, title: str, headers: tuple[str, ...], rows: list[list[object]]):
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    for row in rows:
        worksheet.append([value.label if isinstance(value, Link) else value for value in row])
        row_number = worksheet.max_row
        for column, value in enumerate(row, start=1):
            if isinstance(value, Link):
                cell = worksheet.cell(row=row_number, column=column)
                cell.hyperlink = value.target
                cell.font = LINK_FONT

    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER
    worksheet.row_dimensions[1].height = 32

    for row_number in range(2, worksheet.max_row + 1):
        fill = PatternFill("solid", fgColor=ALT_FILL) if row_number % 2 == 0 else PatternFill(fill_type=None)
        for cell in worksheet[row_number]:
            cell.fill = fill
            if not cell.hyperlink:
                cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = BODY_BORDER

    for column in range(1, worksheet.max_column + 1):
        values = [worksheet.cell(row=row, column=column).value or "" for row in range(1, worksheet.max_row + 1)]
        width = min(60, max(12, max(_display_width(value) for value in values) + 2))
        if headers[column - 1] in {"event_code", "respondent_code", "alter_code", "check_id", "명제", "evidence_status", "disposition"}:
            width = min(width, 20)
        worksheet.column_dimensions[get_column_letter(column)].width = width

    for row_number in range(2, worksheet.max_row + 1):
        line_counts = []
        for column in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=row_number, column=column).value or ""
            width = worksheet.column_dimensions[get_column_letter(column)].width or 12
            line_counts.append(max(1, sum(max(1, (_display_width(line) + int(width) - 1) // int(width)) for line in str(value).splitlines())))
        worksheet.row_dimensions[row_number].height = min(300, max(24, 15 * max(line_counts)))

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=worksheet.max_row, column=worksheet.max_column).coordinate}"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 85
    worksheet.print_area = worksheet.auto_filter.ref
    worksheet.print_title_rows = "1:1"
    dense_print = title in DENSE_PRINT_SHEETS
    worksheet.page_setup.orientation = "landscape" if dense_print or worksheet.max_column > 5 else "portrait"
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A3 if dense_print or worksheet.max_column > 8 else worksheet.PAPERSIZE_A4
    if dense_print:
        worksheet.sheet_properties.pageSetUpPr.fitToPage = False
        worksheet.page_setup.scale = 100
        worksheet.page_setup.fitToWidth = None
        worksheet.page_setup.fitToHeight = None
        worksheet.print_title_cols = PRINT_TITLE_COLUMNS[title]
    else:
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5
    worksheet.oddHeader.center.text = f"&B{title}"
    worksheet.oddFooter.center.text = "&P / &N"
    return worksheet


def _list_validation(worksheet, header: str, options: tuple[str, ...], max_row: int = 500):
    headers = [cell.value for cell in worksheet[1]]
    column_number = headers.index(header) + 1
    letter = get_column_letter(column_number)
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="허용되지 않은 값",
        error="목록에서 값을 선택하십시오.",
    )
    worksheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}{max_row}")
    return validation


def _normalize_xlsx(path: Path) -> None:
    normalized_path = path.with_suffix(".normalized.xlsx")
    fixed_time = (2026, 8, 12, 0, 0, 0)
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(
        normalized_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as target:
        for name in sorted(source.namelist()):
            source_info = source.getinfo(name)
            data = source.read(name)
            if name == "docProps/core.xml":
                data = re.sub(
                    rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2026-08-12T00:00:00Z\g<2>",
                    data,
                )
            info = copy.copy(source_info)
            info.date_time = fixed_time
            target.writestr(info, data)
    os.replace(normalized_path, path)


def build_workbook(output_path: Path) -> None:
    output_path = Path(output_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Doctoral Paper Lab"
    workbook.properties.title = "SW AI change-agent research model v0.3 analysis workbook"
    workbook.properties.description = "Pre-fieldwork analysis templates; P1-P7 are unverified."
    workbook.properties.created = FIXED_TIMESTAMP
    workbook.properties.modified = FIXED_TIMESTAMP

    readme_rows = [
        ["현재 상태", "현장조사 전 연구준비 산출물이다. 인터뷰·모집·이름 생성·조직 자료 수집은 시작하지 않았다.", None],
        ["명제 상태", "P1-P7은 모두 검증 전 후속 연구 명제다.", Link("연구모형 개요", _public_link("downloads/research-design/research-model-v0.3.md"))],
        ["증거 경계", "논문 직접 근거, 통합 해석, 후속 연구 명제 (검증 전)를 구분한다. NotebookLM 산출물은 근거로 사용하지 않는다.", Link("명제 추적표", _public_link("downloads/research-design/proposition-traceability-v0.3.md"))],
        ["입력 원칙", "사건코딩·관계망·부정사례는 승인 뒤 사용할 빈 템플릿이다. 실명·고객명·기밀·계정 정보를 입력하지 않는다.", Link("파일럿 프로토콜·코딩북", _public_link("downloads/research-design/pilot-protocol-and-codingbook-v0.3.md"))],
        ["구성개념 원본", "12개 구성개념의 정의·경계·귀속 규칙", Link("구성개념 사전", _public_link("downloads/research-design/construct-dictionary-v0.3.md"))],
    ]
    _add_sheet(workbook, "README", ("항목", "내용", "source_markdown"), readme_rows)

    construct_headers = (
        "구성개념",
        "분석 수준",
        "정의",
        "포함",
        "제외",
        "관찰 가능한 근거",
        "시간적 위치",
        "가장 가까운 경쟁 구성개념",
        "귀속 규칙",
        "source_markdown",
    )
    _add_sheet(workbook, "구성개념", construct_headers, _construct_rows())

    proposition_headers = (
        "명제",
        "탐색할 관계",
        "evidence_status",
        "직접 근거 출처",
        "통합 해석",
        "분석 수준",
        "자료원",
        "지지 패턴",
        "반증 패턴",
        "경쟁 설명",
        "연구 2 처분",
        "disposition",
        "source_markdown",
    )
    proposition_sheet = _add_sheet(workbook, "명제추적", proposition_headers, _proposition_rows())
    _list_validation(proposition_sheet, "evidence_status", ("검증 전", "수정", "분기", "기각"))
    _list_validation(proposition_sheet, "disposition", ("미검토", "유지", "수정", "분기", "기각"))

    contribution_headers = (
        "논문",
        "evidence_status",
        "직접 근거",
        "통합 해석",
        "연결되는 명제",
        "직접 근거의 한계",
        "후속 연구를 위한 자료 요구",
        "source_markdown",
    )
    contribution_sheet = _add_sheet(workbook, "논문별기여", contribution_headers, _contribution_rows())
    _list_validation(contribution_sheet, "evidence_status", ("논문 직접 근거", "통합 해석", "후속 연구 명제 (검증 전)"))

    event_sheet = _add_sheet(workbook, "사건코딩", EVENT_COLUMNS, [])
    divergence_dimension_validation = DataValidation(
        type="list",
        formula1='"유지,국소 조정,경계 재구성,자료 부족,자료 충돌"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="허용되지 않은 발산성 판정",
        error="유지, 국소 조정, 경계 재구성, 자료 부족 또는 자료 충돌을 선택하십시오.",
    )
    event_sheet.add_data_validation(divergence_dimension_validation)
    for header in ("dv1_decision", "dv2_decision", "dv3_decision", "dv4_decision"):
        column = get_column_letter(EVENT_COLUMNS.index(header) + 1)
        divergence_dimension_validation.add(f"{column}2:{column}500")
    _list_validation(event_sheet, "divergence_event_decision", ("고발산 후보", "저발산 후보", "혼합", "판정 유보"))
    _list_validation(event_sheet, "divergence_aggregation", ("차원 벡터 보존 (합산·평균 금지)",))
    _list_validation(event_sheet, "p6_use_decision", ("경계 연결 평가", "내부 응집·강한 관계 평가", "차원별 분기", "판정 유보"))
    _list_validation(event_sheet, "p7_use_decision", ("고발산 입력 평가", "입력 불충족", "고발산 입력 별도 확인", "판정 유보"))

    relation_sheet = _add_sheet(workbook, "관계망", RELATION_COLUMNS, [])
    relation_validation = DataValidation(
        type="list",
        formula1='"TRUE,FALSE"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="허용되지 않은 관계 값",
        error="TRUE 또는 FALSE를 선택하십시오.",
    )
    relation_sheet.add_data_validation(relation_validation)
    for header in RELATION_COLUMNS[4:8]:
        column = get_column_letter(RELATION_COLUMNS.index(header) + 1)
        relation_validation.add(f"{column}2:{column}500")

    negative_headers = (
        "event_code",
        "source_type",
        "direct_evidence",
        "researcher_interpretation",
        "unexpected_pattern",
        "primary_rival_explanation",
        "follow_up_comparison",
        "disposition",
    )
    negative_sheet = _add_sheet(workbook, "부정사례", negative_headers, [])
    _list_validation(negative_sheet, "disposition", ("미검토", "유지", "수정", "분기", "기각"))

    ethics_headers = ("check_id", "safeguard", "evidence_status", "disposition", "note", "source_markdown")
    ethics_sheet = _add_sheet(workbook, "윤리체크", ethics_headers, _ethics_rows())
    _list_validation(ethics_sheet, "evidence_status", ("미확인", "확인됨", "해당 없음"))
    _list_validation(ethics_sheet, "disposition", ("진행", "보류", "중단"))

    workbook.active = 0
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    _normalize_xlsx(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the v0.3 research-model analysis workbook.")
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    build_workbook(args.output)


if __name__ == "__main__":
    main()
